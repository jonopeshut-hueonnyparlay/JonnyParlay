#!/usr/bin/env python3
"""Regression tests for audit M-4 — CLV reporting in weekly recap.

Before this fix, weekly_recap.py had no CLV section at all. Production logs
show ~21% of picks are captured by the CLV daemon (the rest are "closing_odds
empty" because of daemon downtime, capture-window misses, or missing book
support). Without explicit coverage reporting, a sparse CLV week would be
invisible — the recap would show P/L and ROI but not whether the edge signal
was there.

These tests lock in:
  - compute_clv_summary distinguishes 'not captured' from 'captured = 0'
  - coverage math is correct for empty / partial / full
  - _format_clv_block always renders captured/total before averaged metrics
  - empty-week renders a loud warning, not a neutral number
  - best/worst are resolved from the CAPTURED subset only
  - the xlsx CLV column is blank (not 0.00) for uncaptured rows
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE / "engine"))


# ─────────────────────────────────────────────────────────────────
# _parse_clv — the 'None means missing' invariant
# ─────────────────────────────────────────────────────────────────

def test_parse_clv_blank_is_none():
    from weekly_recap import _parse_clv
    assert _parse_clv("") is None
    assert _parse_clv(None) is None
    assert _parse_clv("   ") is None


def test_parse_clv_zero_is_captured():
    """A stored 0.0 means 'CLV was captured and happened to be neutral'.
    It must NOT collapse to None or the coverage math silently drops it.
    """
    from weekly_recap import _parse_clv
    assert _parse_clv("0.0") == 0.0
    assert _parse_clv("0") == 0.0


def test_parse_clv_unparseable_is_none():
    from weekly_recap import _parse_clv
    assert _parse_clv("n/a") is None
    assert _parse_clv("pending") is None


def test_parse_clv_handles_signed_decimals():
    from weekly_recap import _parse_clv
    assert _parse_clv("0.015") == 0.015
    assert _parse_clv("-0.005") == -0.005


# ─────────────────────────────────────────────────────────────────
# compute_clv_summary — coverage math
# ─────────────────────────────────────────────────────────────────

def test_summary_empty_week():
    from weekly_recap import compute_clv_summary
    s = compute_clv_summary([])
    assert s["total"] == 0
    assert s["captured"] == 0
    assert s["missing"] == 0
    assert s["coverage_pct"] == 0.0
    assert s["avg_clv"] is None
    assert s["beat_close_pct"] is None


def test_summary_all_missing():
    """Every pick has blank clv → 0 captured, 0% coverage, avg is None.

    The key property under test: avg_clv must be None (not 0.0), so the
    Discord block picks the loud warning path.
    """
    from weekly_recap import compute_clv_summary
    picks = [{"clv": ""} for _ in range(5)]
    s = compute_clv_summary(picks)
    assert s["total"] == 5
    assert s["captured"] == 0
    assert s["missing"] == 5
    assert s["coverage_pct"] == 0.0
    assert s["avg_clv"] is None
    assert s["beat_close_pct"] is None
    assert s["best"] is None
    assert s["worst"] is None


def test_summary_partial_coverage():
    """Missing rows are EXCLUDED from the averaged CLV — they don't count as 0."""
    from weekly_recap import compute_clv_summary
    picks = [
        {"clv": "0.02", "player": "A", "stat": "PTS", "direction": "over", "line": "20"},
        {"clv": "",     "player": "B", "stat": "PTS", "direction": "over", "line": "22"},
        {"clv": "-0.01","player": "C", "stat": "AST", "direction": "under", "line": "6.5"},
        {"clv": "",     "player": "D", "stat": "REB", "direction": "over", "line": "10.5"},
    ]
    s = compute_clv_summary(picks)
    assert s["total"] == 4
    assert s["captured"] == 2
    assert s["missing"] == 2
    assert s["coverage_pct"] == 50.0
    # (2.0 + -1.0) / 2 = 0.5 — NOT 0.25 (which would be avg'd over all 4)
    assert s["avg_clv"] == 0.5
    assert s["beat_close"] == 1
    assert s["beat_close_pct"] == 50.0


def test_summary_zero_captured_clv_is_counted():
    """A captured clv of 0.0 counts toward coverage even though it's neutral."""
    from weekly_recap import compute_clv_summary
    picks = [
        {"clv": "0.0",  "player": "A", "stat": "PTS", "direction": "over", "line": "20"},
        {"clv": "0.02", "player": "B", "stat": "PTS", "direction": "over", "line": "25"},
    ]
    s = compute_clv_summary(picks)
    assert s["captured"] == 2
    assert s["coverage_pct"] == 100.0
    assert s["beat_close"] == 1  # only the 0.02 beats close; 0.0 is neutral
    assert s["avg_clv"] == 1.0   # (0 + 2) / 2


def test_summary_best_worst_drawn_from_captured_only():
    """best/worst must reference captured rows — not uncaptured ones."""
    from weekly_recap import compute_clv_summary
    picks = [
        {"clv": "0.03",  "player": "Best",  "stat": "PTS", "direction": "over", "line": "20"},
        {"clv": "",      "player": "NONE",  "stat": "PTS", "direction": "over", "line": "22"},
        {"clv": "-0.02", "player": "Worst", "stat": "AST", "direction": "under", "line": "6"},
    ]
    s = compute_clv_summary(picks)
    assert s["best"] is not None
    assert s["best"][0]["player"] == "Best"
    assert s["best"][1] == 3.0   # 0.03 → 3.0pp
    assert s["worst"][0]["player"] == "Worst"
    assert s["worst"][1] == -2.0


# ─────────────────────────────────────────────────────────────────
# _format_clv_block — presentation contract
# ─────────────────────────────────────────────────────────────────

def test_empty_capture_block_contains_loud_warning():
    """A zero-capture week MUST render a warning emoji and the count."""
    from weekly_recap import _format_clv_block, compute_clv_summary
    block = _format_clv_block(compute_clv_summary([{"clv": ""}] * 3))
    assert "⚠" in block
    assert "No CLV captured" in block
    assert "3/3" in block or "3 picks" in block.lower()
    # Must NOT show a misleading '0.00pp' average.
    assert "+0.00pp" not in block
    assert "-0.00pp" not in block


def test_partial_coverage_block_marks_gap():
    """Partial capture block must call out the missing-count before the avg."""
    from weekly_recap import _format_clv_block, compute_clv_summary
    picks = [
        {"clv": "0.01", "player": "A", "stat": "PTS", "direction": "over", "line": "20"},
        {"clv": "",     "player": "B", "stat": "PTS", "direction": "over", "line": "22"},
    ]
    block = _format_clv_block(compute_clv_summary(picks))
    # Warning marker on partial coverage.
    assert "⚠" in block
    # Coverage ratio rendered.
    assert "1/2" in block
    # The averaged CLV still shown — but the gap is clear.
    assert "Avg CLV" in block
    # Coverage line must appear BEFORE the averaged line in the rendered block.
    cov_idx = block.find("Coverage")
    avg_idx = block.find("Avg CLV")
    assert cov_idx > -1 and avg_idx > -1 and cov_idx < avg_idx


def test_full_coverage_block_has_no_warning():
    from weekly_recap import _format_clv_block, compute_clv_summary
    picks = [
        {"clv": "0.015", "player": "A", "stat": "PTS", "direction": "over", "line": "20"},
        {"clv": "0.010", "player": "B", "stat": "PTS", "direction": "over", "line": "25"},
    ]
    block = _format_clv_block(compute_clv_summary(picks))
    assert "⚠" not in block
    assert "2/2" in block
    assert "Avg CLV" in block
    assert "+1.25pp" in block
    assert "100%" in block  # beat-close rate


# ─────────────────────────────────────────────────────────────────
# Full embed smoke — CLV block is present in the embed description
# ─────────────────────────────────────────────────────────────────

def test_embed_description_includes_clv_block():
    from weekly_recap import build_weekly_embed
    picks = [
        {"date": "2026-04-14", "sport": "NBA", "player": "A", "team": "LAL",
         "stat": "PTS", "direction": "over", "line": "20", "odds": "-110",
         "book": "draftkings", "tier": "T1", "size": "1", "result": "W", "clv": "0.02",
         "run_type": "primary"},
    ]
    payload = build_weekly_embed("2026-04-13", "2026-04-19", picks, picks, suppress_ping=True)
    desc = payload["embeds"][0]["description"]
    assert "CLV" in desc
    assert "edge > everything" in desc


# ─────────────────────────────────────────────────────────────────
# xlsx CLV column — missing rows MUST be blank, not 0
# ─────────────────────────────────────────────────────────────────

@pytest.mark.skipif(
    __import__("importlib").util.find_spec("openpyxl") is None,
    reason="openpyxl not installed",
)
def test_xlsx_clv_column_blank_for_missing_rows():
    """If the daemon never captured closing odds, the xlsx cell must be
    empty — NEVER rendered as 0.00 which would lie about the sample."""
    import io
    import openpyxl
    from weekly_recap import build_weekly_xlsx

    picks = [
        {"date": "2026-04-14", "sport": "NBA", "player": "Captured", "team": "LAL",
         "stat": "PTS", "direction": "over", "line": "20", "odds": "-110",
         "book": "draftkings", "tier": "T1", "size": "1", "result": "W", "clv": "0.02"},
        {"date": "2026-04-15", "sport": "NBA", "player": "Uncaptured", "team": "BOS",
         "stat": "AST", "direction": "under", "line": "5.5", "odds": "-110",
         "book": "fanduel", "tier": "T2", "size": "1", "result": "L", "clv": ""},
    ]
    buf = build_weekly_xlsx(picks, "2026-04-13", "2026-04-19")
    assert buf is not None
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    # Header row
    headers = [c.value for c in ws[1]]
    assert "CLV" in headers
    clv_col = headers.index("CLV") + 1
    # Row 2: captured — clv = 2.0 (pp)
    assert ws.cell(row=2, column=clv_col).value == 2.0
    # Row 3: uncaptured — cell is blank string, NOT 0
    v = ws.cell(row=3, column=clv_col).value
    assert v == "" or v is None, f"Expected blank CLV cell for uncaptured row, got {v!r}"


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v"]))
