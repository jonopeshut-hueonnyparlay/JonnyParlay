"""Export pick_log.csv to a multi-sheet analytical Excel workbook.

Sheets produced:
  - Dashboard:    top-level KPIs (record, profit, ROI, CLV avg)
  - Picks:        every pick, reorganized columns + PnL/running PnL
  - By Tier:      aggregates by tier (T1, T1B, T2, T3, KILLSHOT, ...)
  - By Run Type:  aggregates by run_type (primary, bonus, daily_lay, ...)
  - By Sport:     aggregates by sport (NBA, NHL, MLB)
  - By Stat:      aggregates by stat (PTS, AST, REB, ...)
  - By Book:      aggregates by sportsbook
  - Daily P&L:    one row per date, running totals

Run: python engine/tools/export_pick_log_xlsx.py
     python engine/tools/export_pick_log_xlsx.py --shadow
     python engine/tools/export_pick_log_xlsx.py --in PATH --out PATH
"""
from __future__ import annotations

import argparse
import csv
import sys
from collections import OrderedDict, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_IN  = ROOT / "data" / "pick_log.csv"
DEFAULT_OUT = ROOT / "data" / "pick_log.xlsx"
SHADOW_IN   = ROOT / "data" / "pick_log_custom.csv"
SHADOW_OUT  = ROOT / "data" / "pick_log_custom.xlsx"

# Reorganized column order for the Picks sheet.  Most important left, noisy right.
PICKS_COL_ORDER = [
    "date", "run_time", "sport", "run_type", "tier",
    "player", "team", "stat", "line", "direction",
    "odds", "size",
    "result", "profit", "running_pnl",
    "proj", "win_prob", "edge", "pick_score",
    "book", "game",
    "closing_odds", "clv",
    "is_home", "card_slot", "mode",
    "context_verdict", "context_reason", "context_score",
    "over_p_raw", "legs",
]

STRAIGHT_RUN_TYPES = {"primary", "bonus", "manual"}
PARLAY_RUN_TYPES   = {"sgp", "longshot", "daily_lay"}

PCT_COLS  = {"win_prob", "edge", "clv", "over_p_raw", "context_score", "roi"}
ODDS_COLS = {"odds", "closing_odds"}
NUM2_COLS = {"line", "proj", "pick_score", "size", "profit", "running_pnl",
             "units_risked", "total_profit", "avg_clv", "decimal_odds"}
INT_COLS  = {"n_picks", "n_w", "n_l", "n_p", "n_void"}

# ── Styles ──────────────────────────────────────────────────────────────
HEADER_FONT   = Font(bold=True, color="FFFFFF")
HEADER_FILL   = PatternFill("solid", fgColor="305496")  # navy blue
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center")
TITLE_FONT    = Font(bold=True, size=14, color="305496")
SUBTITLE_FONT = Font(bold=True, size=11, color="305496")

RESULT_FILLS = {
    "W":    PatternFill("solid", fgColor="C6EFCE"),
    "L":    PatternFill("solid", fgColor="FFC7CE"),
    "P":    PatternFill("solid", fgColor="D9D9D9"),
    "VOID": PatternFill("solid", fgColor="D9D9D9"),
}
GREEN_FILL = PatternFill("solid", fgColor="E2EFDA")
RED_FILL   = PatternFill("solid", fgColor="FCE4D6")
THIN       = Side(border_style="thin", color="BFBFBF")
BOX_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── American → decimal odds → profit ────────────────────────────────────

def american_to_decimal(odds: int | float | None) -> float | None:
    if odds is None:
        return None
    o = float(odds)
    if o > 0:
        return 1.0 + o / 100.0
    if o < 0:
        return 1.0 + 100.0 / abs(o)
    return None


def pick_profit(result: str, size: float, odds: float | None) -> float:
    """Profit (in units) for a graded pick.  W = size × (decimal-1), L = -size, P/V/blank = 0."""
    r = (result or "").upper().strip()
    if r == "W":
        d = american_to_decimal(odds)
        if d is None:
            return 0.0
        return round(size * (d - 1.0), 4)
    if r == "L":
        return round(-size, 4)
    return 0.0


# ── CSV loader ──────────────────────────────────────────────────────────

def load_picks(in_path: Path) -> tuple[list[str], list[dict]]:
    with open(in_path, "r", encoding="utf-8", newline="") as f:
        rdr = csv.DictReader(f)
        header = rdr.fieldnames or []
        rows = [{k: _coerce(v, k) for k, v in row.items()} for row in rdr]
    # Sort chronologically (date asc, run_time asc) for running PnL
    rows.sort(key=lambda r: (r.get("date") or "", r.get("run_time") or ""))
    # Compute profit + running PnL
    running = 0.0
    for r in rows:
        result = r.get("result") or ""
        size   = float(r.get("size") or 0.0)
        odds   = r.get("odds")
        prof   = pick_profit(result, size, odds)
        r["profit"]      = prof
        running         += prof
        r["running_pnl"] = round(running, 4)
    return header, rows


def _coerce(raw: Any, colname: str):
    if raw is None or raw == "":
        return None
    if colname == "legs":
        return raw
    s = str(raw)
    try:
        if "." in s or "e" in s.lower():
            return float(s)
        return int(s)
    except ValueError:
        return raw


# ── Sheet builders ──────────────────────────────────────────────────────

def write_picks_sheet(ws, rows: list[dict]) -> None:
    ws.title = "Picks"
    cols = PICKS_COL_ORDER
    # Header
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN

    # Data
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, name in enumerate(cols, start=1):
            val = row.get(name)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            _apply_format(cell, name, val)

    ws.freeze_panes = "F2"
    ws.auto_filter.ref = ws.dimensions
    _autofit(ws, cols, rows[:200])


def _apply_format(cell, colname: str, val: Any) -> None:
    if colname in PCT_COLS and isinstance(val, (int, float)):
        cell.number_format = "0.00%"
    elif colname in ODDS_COLS and isinstance(val, (int, float)):
        cell.number_format = "+#;-#"
    elif colname in NUM2_COLS and isinstance(val, (int, float)):
        cell.number_format = "0.00"
    elif colname in INT_COLS and isinstance(val, (int, float)):
        cell.number_format = "0"

    if colname == "result" and isinstance(val, str):
        f = RESULT_FILLS.get(val.upper())
        if f: cell.fill = f
    elif colname == "clv" and isinstance(val, (int, float)):
        cell.fill = GREEN_FILL if val > 0 else (RED_FILL if val < 0 else PatternFill())
    elif colname in ("profit", "running_pnl") and isinstance(val, (int, float)):
        if val > 0:   cell.fill = GREEN_FILL
        elif val < 0: cell.fill = RED_FILL


def _autofit(ws, cols, sample_rows):
    for i, name in enumerate(cols, start=1):
        max_len = len(str(name))
        for r in sample_rows:
            v = r.get(name)
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 40)


# ── Aggregation helpers ─────────────────────────────────────────────────

def aggregate(rows: list[dict], group_by: str, only_graded: bool = True) -> list[dict]:
    buckets: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        key = (r.get(group_by) or "(blank)") or "(blank)"
        if only_graded and not (r.get("result") or "").upper() in ("W", "L", "P", "VOID"):
            continue
        buckets[str(key)].append(r)

    out = []
    for key, picks in buckets.items():
        n     = len(picks)
        n_w   = sum(1 for p in picks if (p.get("result") or "").upper() == "W")
        n_l   = sum(1 for p in picks if (p.get("result") or "").upper() == "L")
        n_p   = sum(1 for p in picks if (p.get("result") or "").upper() == "P")
        n_v   = sum(1 for p in picks if (p.get("result") or "").upper() == "VOID")
        units = sum(float(p.get("size") or 0.0) for p in picks
                    if (p.get("result") or "").upper() in ("W", "L"))
        prof  = sum(float(p.get("profit") or 0.0) for p in picks)
        roi   = (prof / units) if units > 0 else 0.0
        clvs  = [float(p["clv"]) for p in picks
                 if isinstance(p.get("clv"), (int, float))]
        avg_clv = (sum(clvs) / len(clvs)) if clvs else None
        out.append({
            group_by:       key,
            "n_picks":      n,
            "n_w":          n_w,
            "n_l":          n_l,
            "n_p":          n_p,
            "n_void":       n_v,
            "units_risked": round(units, 2),
            "total_profit": round(prof, 2),
            "roi":          round(roi, 4),
            "avg_clv":      round(avg_clv, 4) if avg_clv is not None else None,
        })
    out.sort(key=lambda r: r["total_profit"] or 0.0, reverse=True)
    return out


def write_agg_sheet(ws, title: str, group_col: str, agg_rows: list[dict]) -> None:
    ws.title = title
    cols = [group_col, "n_picks", "n_w", "n_l", "n_p", "n_void",
            "units_risked", "total_profit", "roi", "avg_clv"]
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
    for r_idx, row in enumerate(agg_rows, start=2):
        for c_idx, name in enumerate(cols, start=1):
            val = row.get(name)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            _apply_format(cell, name, val)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    _autofit(ws, cols, agg_rows)


def write_daily_pnl_sheet(ws, rows: list[dict]) -> None:
    ws.title = "Daily P&L"
    by_date: dict[str, list[dict]] = defaultdict(list)
    for r in rows:
        by_date[r.get("date") or "(blank)"].append(r)
    daily = []
    running = 0.0
    for date in sorted(by_date.keys()):
        picks = by_date[date]
        graded = [p for p in picks if (p.get("result") or "").upper() in ("W", "L", "P", "VOID")]
        units  = sum(float(p.get("size") or 0.0) for p in graded
                     if (p.get("result") or "").upper() in ("W", "L"))
        prof   = sum(float(p.get("profit") or 0.0) for p in picks)
        running += prof
        daily.append({
            "date":         date,
            "n_picks":      len(picks),
            "n_w":          sum(1 for p in graded if (p.get("result") or "").upper() == "W"),
            "n_l":          sum(1 for p in graded if (p.get("result") or "").upper() == "L"),
            "n_p":          sum(1 for p in graded if (p.get("result") or "").upper() == "P"),
            "n_void":       sum(1 for p in graded if (p.get("result") or "").upper() == "VOID"),
            "units_risked": round(units, 2),
            "daily_profit": round(prof, 2),
            "running_pnl":  round(running, 2),
            "roi":          round(prof / units, 4) if units > 0 else 0.0,
        })

    cols = ["date", "n_picks", "n_w", "n_l", "n_p", "n_void",
            "units_risked", "daily_profit", "running_pnl", "roi"]
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN
    for r_idx, row in enumerate(daily, start=2):
        for c_idx, name in enumerate(cols, start=1):
            val = row.get(name)
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            colspec = "profit" if name in ("daily_profit", "running_pnl") else name
            _apply_format(cell, colspec, val)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    _autofit(ws, cols, daily)


def _bet_class_stats(rows: list[dict], run_types: set[str]) -> dict:
    """Aggregate stats for a subset of rows filtered by run_type."""
    subset = [r for r in rows if (r.get("run_type") or "").lower() in run_types]
    graded = [r for r in subset if (r.get("result") or "").upper() in ("W", "L", "P", "VOID")]
    n_w = sum(1 for r in graded if (r.get("result") or "").upper() == "W")
    n_l = sum(1 for r in graded if (r.get("result") or "").upper() == "L")
    n_p = sum(1 for r in graded if (r.get("result") or "").upper() == "P")
    n_v = sum(1 for r in graded if (r.get("result") or "").upper() == "VOID")
    units  = sum(float(r.get("size") or 0.0) for r in graded
                 if (r.get("result") or "").upper() in ("W", "L"))
    profit = sum(float(r.get("profit") or 0.0) for r in subset)
    return {
        "total":    len(subset),
        "graded":   len(graded),
        "pending":  len(subset) - len(graded),
        "n_w":      n_w,
        "n_l":      n_l,
        "n_p":      n_p,
        "n_v":      n_v,
        "win_rate": (n_w / (n_w + n_l)) if (n_w + n_l) > 0 else 0.0,
        "units":    round(units, 2),
        "profit":   round(profit, 2),
        "roi":      (profit / units) if units > 0 else 0.0,
    }


def write_dashboard(ws, rows: list[dict]) -> None:
    ws.title = "Dashboard"
    overall   = _bet_class_stats(rows, STRAIGHT_RUN_TYPES | PARLAY_RUN_TYPES)
    straights = _bet_class_stats(rows, STRAIGHT_RUN_TYPES)
    parlays   = _bet_class_stats(rows, PARLAY_RUN_TYPES)

    clvs = [float(r["clv"]) for r in rows if isinstance(r.get("clv"), (int, float))]
    avg_clv = (sum(clvs) / len(clvs)) if clvs else None
    pos_clv_pct = (sum(1 for c in clvs if c > 0) / len(clvs)) if clvs else None

    ws["A1"] = "JonnyParlay — Pick Log Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    sections = [
        ("Overall (all bet classes)", [
            ("Total picks",             overall["total"]),
            ("Graded",                  overall["graded"]),
            ("Pending / unscored",      overall["pending"]),
            ("Units risked (W+L only)", overall["units"]),
            ("Total profit (units)",    overall["profit"]),
            ("ROI",                     overall["roi"]),
        ]),
        ("Straights — primary / bonus / manual", [
            ("Picks",                   straights["total"]),
            ("Wins / Losses / Push / Void",
                f"{straights['n_w']} / {straights['n_l']} / {straights['n_p']} / {straights['n_v']}"),
            ("Win rate (W / W+L)",      straights["win_rate"]),
            ("Units risked",            straights["units"]),
            ("Total profit (units)",    straights["profit"]),
            ("ROI",                     straights["roi"]),
        ]),
        ("Parlays — sgp / longshot / daily_lay", [
            ("Parlays",                 parlays["total"]),
            ("Wins / Losses / Push / Void",
                f"{parlays['n_w']} / {parlays['n_l']} / {parlays['n_p']} / {parlays['n_v']}"),
            ("Hit rate (W / W+L)",      parlays["win_rate"]),
            ("Units risked",            parlays["units"]),
            ("Total profit (units)",    parlays["profit"]),
            ("ROI",                     parlays["roi"]),
        ]),
        ("CLV", [
            ("Picks with CLV captured", len(clvs)),
            ("Average CLV",             round(avg_clv, 4) if avg_clv is not None else "—"),
            ("Beat the close (%)",      pos_clv_pct if pos_clv_pct is not None else "—"),
        ]),
    ]

    row = 3
    for section_name, items in sections:
        ws.cell(row=row, column=1, value=section_name).font = SUBTITLE_FONT
        row += 1
        for label, val in items:
            ws.cell(row=row, column=1, value=label)
            c = ws.cell(row=row, column=2, value=val)
            c.alignment = Alignment(horizontal="right")
            ll = label.lower()
            if isinstance(val, float):
                if ll.startswith("roi") or "rate" in ll or "(%)" in label or ll.startswith("beat"):
                    c.number_format = "0.00%"
                elif ll.startswith("average clv"):
                    c.number_format = "+0.0000;-0.0000"
                else:
                    c.number_format = "0.00"
            elif isinstance(val, int) and not isinstance(val, bool):
                c.number_format = "0"
            for cc in (ws.cell(row=row, column=1), ws.cell(row=row, column=2)):
                cc.border = BOX_BORDER
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22

    ws.cell(row=row, column=1, value="See other sheets for: Picks (every row), By Tier / Run Type / Sport / Stat / Book, and Daily P&L.").font = Font(italic=True, color="595959")


# ── Main ────────────────────────────────────────────────────────────────

def export(in_path: Path, out_path: Path) -> None:
    if not in_path.exists():
        sys.exit(f"ERROR: {in_path} does not exist")
    header, rows = load_picks(in_path)
    if not rows:
        sys.exit(f"ERROR: {in_path} has no picks (header only)")

    wb = Workbook()
    write_dashboard(wb.active, rows)
    write_picks_sheet(wb.create_sheet("Picks"), rows)
    write_agg_sheet(wb.create_sheet("By Tier"),     "By Tier",     "tier",     aggregate(rows, "tier"))
    write_agg_sheet(wb.create_sheet("By Run Type"), "By Run Type", "run_type", aggregate(rows, "run_type"))
    write_agg_sheet(wb.create_sheet("By Sport"),    "By Sport",    "sport",    aggregate(rows, "sport"))
    write_agg_sheet(wb.create_sheet("By Stat"),     "By Stat",     "stat",     aggregate(rows, "stat"))
    write_agg_sheet(wb.create_sheet("By Book"),     "By Book",     "book",     aggregate(rows, "book"))
    write_daily_pnl_sheet(wb.create_sheet("Daily P&L"), rows)

    try:
        wb.save(out_path)
    except PermissionError:
        # File locked (probably open in Excel). Fall back to timestamped name.
        from datetime import datetime
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = out_path.with_name(f"{out_path.stem}_{stamp}{out_path.suffix}")
        print(f"WARNING: {out_path} is open in Excel — saving to fallback path instead")
        wb.save(fallback)
        out_path = fallback
    print(f"Wrote {out_path}")
    print(f"  {len(rows)} picks, 8 sheets (Dashboard, Picks, 5 aggregates, Daily P&L)")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument("--in",  dest="in_path",  default=None)
    p.add_argument("--out", dest="out_path", default=None)
    p.add_argument("--shadow", action="store_true",
                   help="Use pick_log_custom.csv → pick_log_custom.xlsx")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    if args.shadow:
        in_path  = Path(args.in_path)  if args.in_path  else SHADOW_IN
        out_path = Path(args.out_path) if args.out_path else SHADOW_OUT
    else:
        in_path  = Path(args.in_path)  if args.in_path  else DEFAULT_IN
        out_path = Path(args.out_path) if args.out_path else DEFAULT_OUT
    export(in_path, out_path)


if __name__ == "__main__":
    main()
