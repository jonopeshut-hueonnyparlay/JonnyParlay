#!/usr/bin/env python3
"""
weekly_recap.py — Weekly performance recap for #announcements
Reads pick_log.csv, computes Mon–Sun stats, posts embed + xlsx attachment.
Schedule via Windows Task Scheduler: every Sunday at 9:00 PM ET.

Usage:
    python weekly_recap.py                       # This/most-recent completed week
    python weekly_recap.py --test                # Suppress @everyone
    python weekly_recap.py --week 2026-04-13     # Week containing this date
    python weekly_recap.py --repost              # Force re-post (bypass guard)

Requires: pip install openpyxl --break-system-packages  (for xlsx attachment)
"""

import argparse, csv, io, json, os, sys, time, tempfile
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

# Canonical locked-reader helper — every pick_log reader must take the same
# FileLock as the writers (audit H-8 / M-series).
sys.path.insert(0, str(Path(__file__).resolve().parent))
from pick_log_io import load_rows  # noqa: E402

try:
    import requests
except ImportError:
    print("pip install requests --break-system-packages")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# M9: resolved via paths.py — honours $JONNYPARLAY_ROOT
from paths import (  # noqa: E402
    PICK_LOG_PATH as _PICK_LOG_PATH_P,
    PICK_LOG_MANUAL_PATH as _PICK_LOG_MANUAL_PATH_P,
    DISCORD_GUARD_FILE as _DISCORD_GUARD_FILE_P,
)

# ── Config ────────────────────────────────────────────────────────────────────
PICK_LOG_PATH            = str(_PICK_LOG_PATH_P)
PICK_LOG_MANUAL_PATH     = str(_PICK_LOG_MANUAL_PATH_P)
DISCORD_GUARD_FILE       = str(_DISCORD_GUARD_FILE_P)

# Announce webhook loaded from env/.env — see secrets_config.py (audit C-6).
from secrets_config import DISCORD_ANNOUNCE_WEBHOOK

BRAND_LOGO               = "https://cdn.discordapp.com/attachments/1115840612915228727/1225636209221566625/JonnyParlaylogoRedBlack.png"

# Pick-label formatters moved to pick_labels.py (audit L-3, closed Apr 21
# 2026) so the recap and analyze_picks backtest use the same label source.
# Before this, each file had its own inlined formatter and the PARLAY
# special case only lived here — the analyzer was rendering "3-LEG COVER
# PARLAY" as a prop until someone remembered to fix both sides.
from pick_labels import GAME_LINE_STATS, short_label as _pick_short_label  # noqa: E402

# Only props count toward weekly/monthly totals — parlays excluded from tracking.
# Matches grade_picks.py COUNTED_RUN_TYPES. Manual picks discontinued.
COUNTED_RUN_TYPES = {"primary", "bonus"}

# Sportsbook display contract — canonical definition in book_names.py (audit H-13).
from book_names import BOOK_DISPLAY as _BOOK_DISPLAY, display_book  # noqa: E402

# Locale-independent English month names (audit M-22). ``calendar.month_name``
# returns localized strings on non-en-US Windows installs and leaks foreign
# month names into public Discord posts — picksbyjonny is an English brand.
from month_names import MONTH_NAMES  # noqa: E402

# Brand constants (audit L-7) — central source of truth for the tagline
# so a rename is a single-file edit instead of a multi-file grep sweep.
from brand import BRAND_TAGLINE, BRAND_HANDLE  # noqa: E402

# Shared atomic-JSON writer (architectural note #2) — replaces the inline
# tmp+fsync+replace fallback in _save_guard.
from io_utils import atomic_write_json  # noqa: E402

# ── Core helpers ──────────────────────────────────────────────────────────────

# Result codes that grade_picks.py writes into the ledger.
#   W    — win
#   L    — loss
#   P    — push (bet returned, risked units do not count)
#   VOID — canceled / no-action game. Must be treated EXACTLY like a push:
#          the book refunds the bet, so P&L is 0 AND the risked stake is
#          excluded from ROI. Audit H-5: the previous code silently
#          lumped VOID into "not W, not L, not P" which quietly counted
#          the stake as risked (inflating ROI denominator) while paying
#          out 0 on the numerator.
_REFUNDED_RESULTS = frozenset({"P", "VOID"})


def compute_pl(size, odds_str, result):
    # Audit L-16 (closed Apr 20 2026): this function used to round to 4
    # decimals internally. That's a precision bug waiting to happen — every
    # aggregation (daily_stats, week totals, recap rollups) compounds that
    # rounding error and then the display layer rounds AGAIN to 2 decimals.
    # Return the raw float; format at the presentation boundary only.
    try:
        size = float(size)
        odds = int(float(str(odds_str).replace("+", "")))
    except (ValueError, TypeError):
        return 0.0
    if result == "W":
        return size * (100 / abs(odds)) if odds < 0 else size * (odds / 100)
    elif result == "L":
        return -size
    # P, VOID, "", and any other non-settled code — bet returned / unsettled.
    return 0.0

def daily_stats(picks):
    w      = sum(1 for p in picks if p.get("result") == "W")
    l      = sum(1 for p in picks if p.get("result") == "L")
    # Count pushes AND VOIDs together — they're indistinguishable to ROI math.
    pu     = sum(1 for p in picks if p.get("result") in _REFUNDED_RESULTS)
    pl     = sum(compute_pl(p.get("size", 0), p.get("odds", "-110"), p.get("result", "")) for p in picks)
    # Exclude refunded (P / VOID) stakes from risked. Audit H-5: without the
    # VOID branch, a canceled game's 1u stake inflated the denominator even
    # though the book refunded it, dragging daily ROI lower than reality.
    risked = sum(
        float(p.get("size", 0) or 0) for p in picks
        if p.get("result") not in _REFUNDED_RESULTS
    )
    roi    = (pl / risked * 100) if risked > 0 else 0.0
    return w, l, pu, round(pl, 2), round(roi, 1)


def _parse_clv(raw):
    """Safely parse a stored clv value. Returns None on blank/unparseable.

    Distinguishes 'not captured' (None) from 'captured = 0.0' (float 0.0) —
    this matters because averaging over missing rows as 0 silently inflates
    coverage and masks capture failures (audit M-4).
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def compute_clv_summary(picks):
    """Compute CLV coverage and aggregate metrics for a set of picks.

    A pick is "CLV-captured" when its ``clv`` column has a parseable float.
    Missing closing_odds → blank clv → not captured.

    Returns a dict:
      total         — number of picks considered
      captured      — number with parseable clv
      missing       — total - captured
      coverage_pct  — captured / total * 100 (0 when total=0)
      avg_clv       — mean of captured clv values, as a percentage point
                       (e.g. 0.015 stored → 1.5 returned). None when captured=0.
      beat_close    — count of captured picks with clv > 0
      beat_close_pct — beat_close / captured * 100. None when captured=0.
      best          — (pick, clv_pp) of the best CLV this week, or None
      worst         — (pick, clv_pp) of the worst CLV this week, or None

    Failing loud is the whole point: callers should render ``missing`` and
    ``coverage_pct`` explicitly so a 0% or incomplete capture week doesn't
    look like strong edge.
    """
    total    = len(picks)
    captured = []   # list of (pick, clv_float)
    for p in picks:
        clv = _parse_clv(p.get("clv"))
        if clv is not None:
            captured.append((p, clv))

    cap_n     = len(captured)
    missing   = total - cap_n
    coverage  = (cap_n / total * 100.0) if total > 0 else 0.0

    if cap_n == 0:
        return {
            "total": total, "captured": 0, "missing": missing,
            "coverage_pct": round(coverage, 1),
            "avg_clv": None, "beat_close": 0, "beat_close_pct": None,
            "best": None, "worst": None,
        }

    # avg_clv stored and returned as decimal (e.g. 0.015 = 1.5pp).
    # Multiply by 100 only at final display, not here — keeps scale consistent
    # with clv_report.clv_grade() which also expects decimal.
    clv_vals   = [c for _, c in captured]
    avg        = sum(clv_vals) / len(clv_vals)
    beat       = sum(1 for v in clv_vals if v > 0)
    beat_pct   = beat / cap_n * 100.0
    best_pair  = max(captured, key=lambda pc: pc[1])
    worst_pair = min(captured, key=lambda pc: pc[1])
    return {
        "total": total, "captured": cap_n, "missing": missing,
        "coverage_pct": round(coverage, 1),
        "avg_clv": round(avg, 4),
        "beat_close": beat,
        "beat_close_pct": round(beat_pct, 1),
        "best":  (best_pair[0],  round(best_pair[1]  * 100.0, 2)),
        "worst": (worst_pair[0], round(worst_pair[1] * 100.0, 2)),
    }

def load_picks(path=PICK_LOG_PATH, extra_paths=()):
    """Load rows from main pick log. Manual log excluded — no manual tracking.

    Arch note #3: row-reading goes through ``pick_log_io.load_rows``, which
    takes the shared FileLock per path so concurrent engine/grader/CLV-daemon
    writes can't serve a partial row (audit H-8).
    """
    paths = [path] + list(extra_paths or [])
    return load_rows(paths)

def week_range(ref=None):
    """Return (monday_str, sunday_str) for the most recently completed week.
    If today is Sunday, returns this week (Mon–today).
    """
    today = ref or datetime.now(ZoneInfo("America/New_York"))
    days_since_sunday = (today.weekday() + 1) % 7  # 0 on Sunday
    last_sunday = today - timedelta(days=days_since_sunday)
    last_monday = last_sunday - timedelta(days=6)
    return last_monday.strftime("%Y-%m-%d"), last_sunday.strftime("%Y-%m-%d")

def week_range_containing(ref_date):
    """Return (monday_str, sunday_str) for the week that contains ref_date."""
    monday = ref_date - timedelta(days=ref_date.weekday())
    sunday = monday + timedelta(days=6)
    return monday.strftime("%Y-%m-%d"), sunday.strftime("%Y-%m-%d")

def filter_week(all_rows, mon_str, sun_str):
    # C8: include VOID picks (stake refunded, same as P for record-keeping).
    # Excluding VOID under-counts total bets placed and inflates win-rate.
    return [
        r for r in all_rows
        if r.get("result") in ("W", "L", "P", "VOID")
        and r.get("run_type") in COUNTED_RUN_TYPES
        and mon_str <= r.get("date", "") <= sun_str
    ]

def _fmt_week_label(mon_str, sun_str):
    """'Apr 7 – Apr 13'  (cross-platform, no leading zero)"""
    mon_dt = datetime.strptime(mon_str, "%Y-%m-%d")
    sun_dt = datetime.strptime(sun_str, "%Y-%m-%d")
    return f"{mon_dt.strftime('%b')} {mon_dt.day} – {sun_dt.strftime('%b')} {sun_dt.day}"

# ── xlsx builder ──────────────────────────────────────────────────────────────

#: Audit H-15 — hard cap on rows materialized in the weekly xlsx BytesIO.
#: A normal week is 30-80 picks; this ceiling only triggers on backfill /
#: replay jobs that would otherwise OOM the process. Picked at 5000 because
#: that's a full year of the current cadence with headroom. If this ever
#: trips in real use, it's a signal the caller is passing the wrong slice.
WEEKLY_XLSX_ROW_CAP: int = 5000


def build_weekly_xlsx(week_picks, mon_str, sun_str):
    """Build xlsx. Returns BytesIO or None if openpyxl not installed.

    Audit H-15: the whole workbook is built in a BytesIO. For a normal week
    (<100 picks) that's fine, but a misrouted backfill / replay could hand
    this function every pick ever logged and OOM the process. We cap the
    row count at :data:`WEEKLY_XLSX_ROW_CAP` and emit a warning; callers
    that actually want a full-history dump should stream row-by-row
    instead of calling this helper.
    """
    if not _HAS_OPENPYXL:
        return None

    # H-15: cap the row count. We truncate (most-recent first) rather than
    # refusing to build anything — the user still gets a usable file and a
    # warning telling them their slice was too large.
    if len(week_picks) > WEEKLY_XLSX_ROW_CAP:
        print(
            f"  [weekly-xlsx] ⚠ H-15: week_picks has {len(week_picks)} rows, "
            f"capping at WEEKLY_XLSX_ROW_CAP={WEEKLY_XLSX_ROW_CAP}. This helper "
            f"is meant for a single week (<100 rows); if you need more, stream "
            f"rows directly instead of materializing into BytesIO."
        )
        # Keep the most recent rows — prefer the end of the list since
        # callers typically append chronologically.
        week_picks = week_picks[-WEEKLY_XLSX_ROW_CAP:]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Picks"

    dark_fill = PatternFill("solid", fgColor="111118")
    alt_fill  = PatternFill("solid", fgColor="16161F")
    hdr_fill  = PatternFill("solid", fgColor="1a1a28")
    hdr_font  = Font(bold=True, color="FFD700", size=11)
    row_bdr   = Border(bottom=Side(style="thin", color="333344"))
    center    = Alignment(horizontal="center", vertical="center")

    # Column layout (audit M-4): CLV is the 13th column, between Result and P/L.
    # Blank cells in the CLV column mean "closing_odds never captured" — do
    # NOT render them as 0.00, or the xlsx will silently average missing data
    # down to neutral.
    headers = ["Date", "Sport", "Pick", "Stat", "Dir", "Line", "Odds", "Book",
               "Tier", "Size", "Result", "CLV", "P/L"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment = hdr_font, hdr_fill, center
    ws.row_dimensions[1].height = 22

    for ri, p in enumerate(week_picks, 2):
        pl     = compute_pl(p.get("size", 0), p.get("odds", "-110"), p.get("result", ""))
        result = p.get("result", "")
        fill   = alt_fill if ri % 2 == 0 else dark_fill
        clv_raw = _parse_clv(p.get("clv"))
        clv_val = round(clv_raw * 100.0, 2) if clv_raw is not None else ""
        values = [
            p.get("date", ""), p.get("sport", ""), _pick_short_label(p),
            p.get("stat", ""), p.get("direction", "").upper(), p.get("line", ""),
            p.get("odds", ""), display_book(p.get("book", "")),
            p.get("tier", ""), float(p.get("size", 0) or 0),
            result, clv_val, round(pl, 2),
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.alignment, c.border, c.fill = center, row_bdr, fill
            if col == 11:  # Result
                c.font = Font(color=("2ECC71" if result == "W" else "FF4444" if result == "L" else "AAAAAA"), bold=True)
            if col == 12:  # CLV — green if beat close, red if lost CLV, grey on blank
                if clv_raw is None:
                    c.font = Font(color="666677", italic=True)
                elif clv_raw > 0:
                    c.font = Font(color="2ECC71", bold=True)
                elif clv_raw < 0:
                    c.font = Font(color="FF4444", bold=True)
                else:
                    c.font = Font(color="AAAAAA")
            if col == 13:  # P/L
                c.font = Font(color="2ECC71" if pl >= 0 else "FF4444", bold=True)
        ws.row_dimensions[ri].height = 18

    # Totals
    tr = len(week_picks) + 2
    w, l, pu, pl_total, roi = daily_stats(week_picks)
    clv_summary = compute_clv_summary(week_picks)
    ws.cell(row=tr, column=1,  value="TOTALS").font  = Font(bold=True, color="FFD700", size=11)
    ws.cell(row=tr, column=11, value=f"{w}W-{l}L{('-' + str(pu) + 'P') if pu else ''}").font = Font(bold=True, color="FFD700")
    if clv_summary["avg_clv"] is not None:
        avg_label = f"{clv_summary['avg_clv'] * 100:+.2f}pp ({clv_summary['captured']}/{clv_summary['total']})"
        ws.cell(row=tr, column=12, value=avg_label).font = Font(
            bold=True, color="2ECC71" if clv_summary['avg_clv'] >= 0 else "FF4444"
        )
    else:
        ws.cell(row=tr, column=12, value="no CLV").font = Font(bold=True, color="FF4444", italic=True)
    ws.cell(row=tr, column=13, value=round(pl_total, 2)).font = Font(bold=True, color="2ECC71" if pl_total >= 0 else "FF4444")
    ws.cell(row=tr + 1, column=13, value=f"ROI: {roi:+.1f}%").font = Font(color="FFD700", bold=True)

    # Column widths: added CLV column (width 10) before P/L.
    for col, w_ in enumerate([12, 6, 36, 8, 5, 6, 8, 16, 10, 6, 8, 10, 8], 1):
        ws.column_dimensions[get_column_letter(col)].width = w_
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Discord embed ─────────────────────────────────────────────────────────────


def _format_clv_block(summary):
    """Render the CLV summary as a Discord-safe multi-line string.

    Three cases:
      1. zero captured → loud warning: "CLV: ⚠ 0/N captured"
      2. partial capture (coverage < 100%) → show metrics + explicit gap
         so the reader knows the average is over a subset
      3. full capture → clean metrics block

    Always puts the coverage count first so it's impossible to read the
    averaged CLV without also seeing how many picks it's averaged over.
    """
    cap   = summary["captured"]
    total = summary["total"]
    miss  = summary["missing"]

    header = "━━━━━━━━━━━━━━━━"

    if cap == 0:
        body = [
            f"📊 **CLV** ({BRAND_TAGLINE})",
            f"⚠ No CLV captured this week — {miss}/{total} picks missing closing odds.",
            "_Check `data/clv_daemon.log` — capture gap may indicate daemon downtime._",
        ]
        return "\n".join([header] + body)

    avg_str = f"{summary['avg_clv'] * 100:+.2f}pp"
    beat_str = f"{summary['beat_close']}/{cap} ({summary['beat_close_pct']:.0f}%)"

    body = [f"📊 **CLV**"]
    if miss > 0:
        # Partial capture — mark it plainly so the averaged figure isn't
        # read as a full-week number.
        body.append(
            f"⚠ Coverage: **{cap}/{total}** captured | "
            f"**{miss}** missing closing odds"
        )
    else:
        body.append(f"Coverage: {cap}/{total} captured")

    body.append(f"Avg CLV: **{avg_str}** | Beat close: **{beat_str}**")

    if summary.get("best"):
        best_p, best_v = summary["best"]
        body.append(f"🔼 Best CLV: {_pick_short_label(best_p)} | {best_v:+.2f}pp")
    if summary.get("worst") and summary["worst"] != summary.get("best"):
        worst_p, worst_v = summary["worst"]
        body.append(f"🔽 Worst CLV: {_pick_short_label(worst_p)} | {worst_v:+.2f}pp")

    return "\n".join([header] + body)


def _fmt_clock_et(dt) -> str:
    """Format ET clock time as 'H:MM AM/PM ET' without leading zero and
    without locale sensitivity.

    Audit M-17 (closed Apr 21 2026): the old ``strftime("%I:%M %p ET")``
    produced "08:35 PM ET" on Windows (leading zero) and "08:35 nachm. ET"
    on a German locale (localized %p). The rest of the codebase explicitly
    avoids ``%d`` for the same leading-zero reason, so this helper brings
    the clock line in line with that house style — and builds AM/PM in
    English regardless of the system locale.
    """
    hour12 = dt.hour % 12 or 12
    ampm = "AM" if dt.hour < 12 else "PM"
    return f"{hour12}:{dt.minute:02d} {ampm} ET"


def build_weekly_embed(mon_str, sun_str, week_picks, all_rows, suppress_ping=False):
    w, l, pu, pl, roi = daily_stats(week_picks)
    pl_str  = f"+{pl:.2f}u" if pl >= 0 else f"{pl:.2f}u"
    roi_str = f"+{roi:.1f}%" if roi >= 0 else f"{roi:.1f}%"
    week_label = _fmt_week_label(mon_str, sun_str)
    now_str    = _fmt_clock_et(datetime.now(ZoneInfo("America/New_York")))

    # Daily breakdown
    grouped = defaultdict(list)
    for p in week_picks:
        grouped[p["date"]].append(p)
    day_lines = []
    for d in sorted(grouped.keys()):
        dw, dl, _, dpl, _ = daily_stats(grouped[d])
        dpl_str   = f"+{dpl:.2f}u" if dpl >= 0 else f"{dpl:.2f}u"
        emoji     = "✅" if dpl > 0 else ("❌" if dpl < 0 else "➖")
        dt        = datetime.strptime(d, "%Y-%m-%d")
        day_label = f"{dt.strftime('%a')} {dt.day}"
        day_lines.append(f"{emoji} **{day_label}:** {dw}-{dl} | {dpl_str}")

    # Tier breakdown removed from public embed — internal diagnostic only (analyze_picks.py)

    # Best / worst
    pick_pls = [(p, compute_pl(p.get("size",0), p.get("odds","-110"), p.get("result",""))) for p in week_picks]
    best  = max(pick_pls, key=lambda x: x[1], default=None)
    worst = min(pick_pls, key=lambda x: x[1], default=None)
    best_line  = f"\n🏆 **Best:** {_pick_short_label(best[0])} | {best[1]:+.2f}u"   if best  else ""
    worst_line = f"\n💀 **Worst:** {_pick_short_label(worst[0])} | {worst[1]:+.2f}u" if (worst and worst != best) else ""

    # Month running total
    dt = datetime.strptime(sun_str, "%Y-%m-%d")
    month_prefix = f"{dt.year}-{dt.month:02d}-"
    month_picks  = [r for r in all_rows
                    if r.get("result") in ("W","L","P","VOID")  # L5: VOID grouped with P
                    and r.get("run_type") in COUNTED_RUN_TYPES
                    and r.get("date","").startswith(month_prefix)]
    mw, ml, _, mpl, _ = daily_stats(month_picks)
    mpl_str = f"+{mpl:.1f}u" if mpl >= 0 else f"{mpl:.1f}u"

    # CLV summary — primary edge indicator. Rendered loud so a capture gap
    # (missing closing_odds from the CLV daemon) doesn't silently inflate
    # the edge signal (audit M-4).
    clv_block = _format_clv_block(compute_clv_summary(week_picks))

    has_xlsx   = _HAS_OPENPYXL
    footer_cta = "Full breakdown attached ↓" if has_xlsx else BRAND_TAGLINE
    color      = 0xFFD700 if pl >= 0 else 0xFF4444
    content    = "" if suppress_ping else "@everyone"

    desc = "\n".join([
        f"**{w}-{l} ({round(w/(w+l)*100) if w+l else 0}%) | {pl_str} | ROI {roi_str}**",
        "",
        "\n".join(day_lines),
        "",
        "━━━━━━━━━━━━━━━━",
        best_line + worst_line,
        "",
        clv_block,
        "",
        f"**{MONTH_NAMES[dt.month]} so far:** {mw}-{ml} | {mpl_str}",
        "",
        "━━━━━━━━━━━━━━━━",
        footer_cta,
    ]).strip()

    return {
        "username": "PicksByJonny",
        "content":  content,
        "embeds": [{
            "title":       f"📅 Weekly Recap — {week_label}",
            "description": desc,
            "color":       color,
            "thumbnail":   {"url": BRAND_LOGO},
            "footer":      {"text": f"{BRAND_HANDLE} | {BRAND_TAGLINE} | {now_str}"},
        }]
    }


# ── Guard helpers ─────────────────────────────────────────────────────────────

_GUARD_TTL_DAYS = 90


try:
    from discord_guard import (
        load_guard as _shared_load_guard,
        save_guard as _shared_save_guard,
        claim_post as _shared_claim_post,
        release_post as _shared_release_post,
        prune_guard as _shared_prune_guard,
    )
    _HAS_SHARED_GUARD = True
except ImportError:
    _HAS_SHARED_GUARD = False


def _prune_guard(guard):
    cutoff = datetime.now(ZoneInfo("America/New_York")).replace(tzinfo=None) - timedelta(days=_GUARD_TTL_DAYS)
    pruned = {}
    for key, val in guard.items():
        keep = True
        for p in key.split(":"):
            if len(p) == 10 and p[4] == "-" and p[7] == "-":
                try:
                    dt = datetime.strptime(p, "%Y-%m-%d")
                    if dt < cutoff:
                        keep = False
                    break
                except ValueError:
                    continue
        if keep:
            pruned[key] = val
    return pruned


def _load_guard():
    # Delegate to shared cross-process-safe helper if available
    if _HAS_SHARED_GUARD:
        return _shared_load_guard()
    try:
        with open(DISCORD_GUARD_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def _save_guard(guard):
    # Delegate to shared cross-process-safe helper if available
    if _HAS_SHARED_GUARD:
        _shared_save_guard(guard)
        return
    try:
        atomic_write_json(DISCORD_GUARD_FILE, _prune_guard(guard))
    except Exception as e:
        # H23: do NOT fall back to a direct open() write — a crash or flush
        # failure mid-write would corrupt the guard file (truncate to zero).
        # Accept that the guard isn't persisted; worst case is a re-post on
        # the next run (use --repost to intentionally force one).
        import logging as _lg
        _lg.getLogger("weekly_recap").warning(
            "Guard save failed — guard not persisted (use --repost to prevent re-post): %s", e
        )


# ── Webhook post with optional file ──────────────────────────────────────────

def _webhook_post_with_file(url, payload, file_buf=None, filename="weekly.xlsx"):
    """POST to Discord webhook, optionally with an xlsx file attachment.
    Retries on 429 (rate limit) and transient 5xx errors; fails fast on 4xx.
    """
    if not url:
        return False
    for attempt in range(1, 4):
        try:
            if file_buf is not None:
                file_buf.seek(0)
                data  = {"payload_json": json.dumps(payload)}
                files = {"files[0]": (
                    filename, file_buf,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )}
                r = requests.post(url, data=data, files=files, timeout=20)
            else:
                r = requests.post(url, json=payload, timeout=10)
            if r.status_code == 429:
                time.sleep(float(r.json().get("retry_after", 2.0)))
                continue
            if r.status_code in (200, 204):
                return True
            # Retry transient 5xx; fail fast on 4xx
            if 500 <= r.status_code < 600 and attempt < 3:
                print(f"  ⚠ Discord 5xx {r.status_code} (attempt {attempt}/3) — retrying")
                time.sleep(2 ** attempt)
                continue
            print(f"  ⚠ Discord {r.status_code}: {r.text[:200]}")
            return False
        except requests.exceptions.RequestException as e:
            if attempt < 3:
                print(f"  ⚠ Discord transport error (attempt {attempt}/3): {e} — retrying")
                time.sleep(2 ** attempt)
            else:
                print(f"  ⚠ Discord post error: {e}")
    return False


# ── Main post function ────────────────────────────────────────────────────────

def post_weekly_recap(week_picks, mon_str, sun_str, all_rows, suppress_ping=False, force=False):
    """Build and post the weekly recap embed + xlsx to #announcements."""
    if not week_picks:
        print(f"  No graded picks for week {mon_str} – {sun_str}")
        return False

    guard_key = f"weekly:{mon_str}"

    if not force:
        if _HAS_SHARED_GUARD:
            if not _shared_claim_post(guard_key):
                print(f"  [Discord] ⏭️  Weekly recap already posted for {mon_str} — use --repost to override")
                return False
        else:
            guard = _load_guard()
            if guard.get(guard_key):
                print(f"  [Discord] ⏭️  Weekly recap already posted for {mon_str} — use --repost to override")
                return False
            guard[guard_key] = True  # pre-claim in fallback path
    else:
        # force=True (--repost): mark claimed so we can release on failure
        if _HAS_SHARED_GUARD:
            _shared_claim_post(guard_key)  # may already be claimed; that's OK

    payload   = build_weekly_embed(mon_str, sun_str, week_picks, all_rows, suppress_ping=suppress_ping)
    xlsx_buf  = build_weekly_xlsx(week_picks, mon_str, sun_str)
    xlsx_name = f"picks_week_{mon_str}.xlsx"

    if not _HAS_OPENPYXL:
        print("  ⚠  openpyxl not installed — no xlsx attachment. Run: pip install openpyxl --break-system-packages")

    if _webhook_post_with_file(DISCORD_ANNOUNCE_WEBHOOK, payload, xlsx_buf, xlsx_name):
        w, l, _, pl, roi = daily_stats(week_picks)
        pl_str = f"+{pl:.2f}u" if pl >= 0 else f"{pl:.2f}u"
        print(f"  [Discord] ✅ Weekly recap posted — {w}W-{l}L |{pl_str} |week of {_fmt_week_label(mon_str, sun_str)}")
        if not _HAS_SHARED_GUARD:
            _save_guard(guard)  # fallback: persist the pre-claimed guard
        return True

    # Webhook failed — release the claim so a retry can re-post
    if _HAS_SHARED_GUARD:
        _shared_release_post(guard_key)
    print("  [Discord] ❌ Weekly recap post failed")
    return False


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Post weekly recap to #announcements")
    parser.add_argument("--week",   default=None,
                        help="Any date within the target week (YYYY-MM-DD). "
                             "Defaults to most recently completed week.")
    parser.add_argument("--test",   action="store_true", help="Suppress @everyone ping")
    parser.add_argument("--repost", action="store_true", help="Force re-post (bypass guard)")
    args = parser.parse_args()

    if args.week:
        ref_date = datetime.strptime(args.week, "%Y-%m-%d")
        mon_str, sun_str = week_range_containing(ref_date)
    else:
        mon_str, sun_str = week_range()

    week_label = _fmt_week_label(mon_str, sun_str)
    print(f"\nWeekly recap: {week_label}  ({mon_str} – {sun_str})")

    all_rows = load_picks()
    if not all_rows:
        print("  ❌ pick_log.csv not found")
        sys.exit(1)

    week_picks = filter_week(all_rows, mon_str, sun_str)
    if not week_picks:
        print(f"  ❌ No graded picks for {week_label}")
        sys.exit(0)

    w, l, pu, pl, roi = daily_stats(week_picks)
    pl_str = f"+{pl:.2f}u" if pl >= 0 else f"{pl:.2f}u"
    print(f"  {w}W-{l}L |{pl_str} |ROI {roi:+.1f}%  ({len(week_picks)} picks)")

    posted = post_weekly_recap(
        week_picks, mon_str, sun_str, all_rows,
        suppress_ping=args.test,
        force=args.repost,
    )
    # AUDIT H-7: if the Discord post failed (network, 4xx, Cloudflare), exit
    # non-zero so Task Scheduler surfaces the failure instead of silently
    # swallowing it. Returning 0 on a failed post makes outages invisible.
    if not posted:
        print("  [weekly-recap] ❌ H-7: post failed — exiting 2 so the scheduler flags it.")
        # Fire the secondary fallback webhook alert (audit H-7, closed Apr 20
        # 2026). No-op if DISCORD_FALLBACK_WEBHOOK isn't configured, and
        # swallows its own failures — it must not mask the real exit code.
        try:
            from webhook_fallback import notify_fallback
            notify_fallback("weekly_recap", err=f"week={mon_str}_to_{sun_str}")
        except Exception as _e:  # noqa: BLE001
            print(f"  [weekly-recap] fallback notifier raised (suppressed): {_e}")
        sys.exit(2)


if __name__ == "__main__":
    main()
